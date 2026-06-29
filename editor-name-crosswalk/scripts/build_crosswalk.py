"""
Build editor name crosswalk: maps EIC canonical names to all name variants
found in the paper database (期刊文章output.xlsx).

Usage:
    python build_crosswalk.py --eic-xlsx 期刊目录.xlsx --paper-xlsx 期刊文章output.xlsx --output 主编名称对照.xlsx
"""
import sys
import re
import openpyxl
from collections import defaultdict


def read_eic_names(eic_xlsx_path):
    """Read all unique EIC names from 期刊目录.xlsx EIC_Long sheet."""
    names = set()
    wb = openpyxl.load_workbook(eic_xlsx_path)
    if "EIC_Long" in wb.sheetnames:
        ws = wb["EIC_Long"]
        for r in range(2, ws.max_row + 1):
            name = ws.cell(row=r, column=6).value  # Col F = 姓名
            if name and str(name).strip() != "None":
                names.add(str(name).strip())
    else:
        # Fallback: read from Sheet1 old format
        ws = wb["Sheet1"]
        for r in range(2, ws.max_row + 1):
            for col in range(6, 50, 7):  # Old EIC layout
                name = ws.cell(row=r, column=col).value
                if name and str(name).strip() != "None":
                    names.add(str(name).strip())
    
    # Also read from Sheet1 metadata (migrated cols)
    print(f"Found {len(names)} unique EIC names")
    return sorted(names)


def read_author_names(paper_xlsx_path):
    """Read all unique author names from paper database.
    Supports both old format (期刊文章output.xlsx with Author 1-10 columns)
    and new format (WoS_Papers_All.xlsx with Authors sheet).
    """
    wb = openpyxl.load_workbook(paper_xlsx_path, read_only=True)
    
    author_forms = defaultdict(set)
    all_raw_names = set()
    
    # Detect format: check if "Authors" sheet exists
    if "Authors" in wb.sheetnames:
        # ── New format: Authors sheet with author_name column ──
        ws = wb["Authors"]
        # Find author_name column index from header
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        name_col = None
        for i, h in enumerate(header):
            if h and str(h).lower().strip() == "author_name":
                name_col = i
                break
        if name_col is None:
            print("  ERROR: author_name column not found in Authors sheet")
            return author_forms
        
        for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if r_idx % 5000 == 0:
                print(f"  Processing row {r_idx}...")
            if name_col < len(row) and row[name_col]:
                name = str(row[name_col]).strip()
                if name and name != "None":
                    all_raw_names.add(name)
                    key = normalize_name_key(name)
                    author_forms[key].add(name)
    else:
        # ── Old format: Sheet1 with Author 1-10 + Corresponding Author ──
        ws = wb[wb.sheetnames[0]]
        for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if r_idx % 2000 == 0:
                print(f"  Processing row {r_idx}...")
            for col_idx in range(2, 12):
                if col_idx < len(row) and row[col_idx]:
                    name = str(row[col_idx]).strip()
                    if name and name != "None":
                        all_raw_names.add(name)
                        key = normalize_name_key(name)
                        author_forms[key].add(name)
            if len(row) > 16 and row[16]:
                name = str(row[16]).strip()
                if name and name != "None":
                    all_raw_names.add(name)
                    key = normalize_name_key(name)
                    author_forms[key].add(name)
    
    print(f"Total unique author name forms: {len(all_raw_names)}")
    print(f"Total unique author keys: {len(author_forms)}")
    return author_forms


def normalize_name_key(name):
    """Normalize a name to a matching key: lowercase, strip punctuation, extract surname."""
    name = name.lower().strip()
    # Remove common patterns like "Jr.", "III", "II"
    name = re.sub(r'\b(jr\.?|sr\.?|iii|ii|iv)\b', '', name)
    # Split by comma: "Luttmer, Erzo F. P." -> surname = "luttmer"
    if ',' in name:
        surname = name.split(',')[0].strip()
        given = name.split(',')[1].strip() if ',' in name else ''
    else:
        parts = name.split()
        surname = parts[-1] if parts else name
        given = ' '.join(parts[:-1])
    
    # Remove diacritics for matching
    surname = surname.replace('ä', 'a').replace('ö', 'o').replace('ü', 'u')
    surname = surname.replace('é', 'e').replace('è', 'e').replace('ê', 'e')
    surname = surname.replace('í', 'i').replace('ó', 'o').replace('ú', 'u')
    surname = surname.replace('ñ', 'n')
    surname = surname.replace('\u0308', '')  # combining diaeresis
    surname = re.sub(r'[^a-z]', '', surname)
    
    # Extract initials from given name
    # Handle both "Erzo F. P." (words) and "EFP" (consecutive uppercase)
    given_clean = re.sub(r'[^a-z\s]', '', given.lower())
    initials = []
    for part in given_clean.split():
        if len(part) <= 3 and part.isalpha():
            # Short parts like "efp" -> treat each letter as initial
            for ch in part:
                initials.append(ch)
        else:
            # Normal words -> first letter is initial
            if part and part[0].isalpha():
                initials.append(part[0])
    
    return f"{surname}|{''.join(initials)}"


def generate_name_variants(canonical_name):
    """Generate all plausible name variants for an EIC name."""
    variants = set()
    variants.add(canonical_name)
    
    parts = canonical_name.split()
    
    # Extract surname (last part) and given names
    if len(parts) >= 2:
        surname = parts[-1]
        given = parts[:-1]  # Could include middle initials
        
        # "Lastname, Firstname" format
        first_names = [p for p in given if len(p) > 1 or not p.endswith('.')]
        initials = [p.rstrip('.') for p in given]
        
        # Variant 1: Lastname, Firstname
        if first_names:
            variants.add(f"{surname}, {' '.join(first_names)}")
        
        # Variant 2: Lastname, FirstInitial
        if initials:
            variants.add(f"{surname}, {''.join(initials)}")
            variants.add(f"{surname}, {' '.join(initials)}")
        
        # Variant 3: Lastname, FirstInitial only (single letter)
        if initials:
            variants.add(f"{surname}, {initials[0]}")
        
        # Variant 4: Handle hyphenated surnames
        if '-' in surname:
            parts_surname = surname.split('-')
            for ps in parts_surname:
                if first_names:
                    variants.add(f"{ps}, {' '.join(first_names)}")
                if initials:
                    variants.add(f"{ps}, {initials[0]}")
        
        # Variant 5: With middle initial variations
        for i in range(len(initials)):
            if i > 0:
                variants.add(f"{surname}, {''.join(initials[:i+1])}")
                variants.add(f"{surname}, {' '.join(initials[:i+1])}")
    
    return variants


def find_matches(eic_names, author_forms):
    """For each EIC, find matching author name forms in the paper database."""
    results = {}
    
    for eic_name in eic_names:
        variants = generate_name_variants(eic_name)
        matched = set()
        
        for variant in variants:
            key = normalize_name_key(variant)
            if key in author_forms:
                matched.update(author_forms[key])
        
        # Also try direct matching: search for any author key starting with the same surname
        eic_key = normalize_name_key(eic_name)
        eic_surname = eic_key.split('|')[0]
        eic_initials_full = eic_key.split('|')[1] if '|' in eic_key else ''
        for a_key, a_forms in author_forms.items():
            a_surname = a_key.split('|')[0]
            a_initials_full = a_key.split('|')[1] if '|' in a_key else ''
            if a_surname != eic_surname or len(eic_surname) < 3:
                continue
            if not eic_initials_full or not a_initials_full:
                continue
            # Require strict initials match to avoid false positives
            matched_initials = (eic_initials_full.startswith(a_initials_full) or 
                               a_initials_full.startswith(eic_initials_full))
            if matched_initials:
                for form in a_forms:
                    if _is_plausible_match(eic_name, form):
                        matched.add(form)
        
        if matched:
            results[eic_name] = sorted(matched)
        else:
            results[eic_name] = []
    
    return results


def _is_plausible_match(eic_name, author_form):
    """Check if an author name form is plausibly the same person as the EIC."""
    eic_parts = eic_name.lower().split()
    if ',' in author_form:
        surname = author_form.split(',')[0].strip().lower()
        given = author_form.split(',')[1].strip().lower() if ',' in author_form else ''
    else:
        parts = author_form.split()
        surname = parts[-1].strip().lower() if parts else ''
        given = ' '.join(parts[:-1]).strip().lower() if len(parts) > 1 else ''
    
    eic_surname = eic_parts[-1].lower().replace('-', '')
    if surname.replace('-', '') != eic_surname:
        return False
    
    # Given name match: initials must be consistent.
    # If both sides have 2+ initials, require at least the second letter
    # to also match (to avoid EFP vs EGJ false positives).
    eic_given = eic_parts[:-1] if len(eic_parts) > 1 else []
    eic_initials = ''.join(p[0] for p in eic_given).lower()
    given_initials = ''.join(p[0] for p in given.split()).lower()
    
    if not eic_initials or not given_initials:
        return True
    
    if len(eic_initials) == 1 and len(given_initials) == 1:
        return eic_initials == given_initials
    
    # Both have multiple initials: require ALL eic_initials to match
    # the start of given_initials, or vice versa (handles F.P. vs FP)
    if len(eic_initials) >= 2 and len(given_initials) >= 2:
        return (eic_initials.startswith(given_initials) or 
                given_initials.startswith(eic_initials))
    
    # One has 1 init, other has multiple: first letter match is enough
    return eic_initials[0] == given_initials[0]


def write_crosswalk(output_path, results):
    """Write the crosswalk to Excel.
    Format:
      Col A: 主编规范名 (canonical EIC name)
      Col B: 数据库中匹配到的名称变体 (pipe-separated list of all matched variants)
      Col C: 匹配数量
      Col D: 匹配状态 (已匹配 / 未找到匹配)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Name Crosswalk"
    
    ws.cell(row=1, column=1).value = "主编规范名"
    ws.cell(row=1, column=2).value = "数据库中匹配到的名称变体"
    ws.cell(row=1, column=3).value = "匹配数量"
    ws.cell(row=1, column=4).value = "匹配状态"
    
    row = 2
    for eic_name, matched in sorted(results.items()):
        ws.cell(row=row, column=1).value = eic_name
        if matched:
            ws.cell(row=row, column=2).value = "|".join(matched)
            ws.cell(row=row, column=3).value = len(matched)
            ws.cell(row=row, column=4).value = "已匹配"
        else:
            ws.cell(row=row, column=3).value = 0
            ws.cell(row=row, column=4).value = "未找到匹配"
        
        row += 1
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    
    wb.save(output_path)
    
    matched_count = sum(1 for v in results.values() if v)
    total = len(results)
    print(f"\nOutput: {output_path}")
    print(f"Matched: {matched_count}/{total} editors")
    print(f"Total name variants found: {sum(len(v) for v in results.values())}")


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Build editor name crosswalk")
    parser.add_argument("--eic-xlsx", required=True, help="Path to 期刊目录.xlsx")
    parser.add_argument("--paper-xlsx", required=True, help="Path to 期刊文章output.xlsx")
    parser.add_argument("--output", required=True, help="Output Excel path")
    args = parser.parse_args()
    
    print("Reading EIC names...")
    eic_names = read_eic_names(args.eic_xlsx)
    
    print("\nReading author names from paper database...")
    author_forms = read_author_names(args.paper_xlsx)
    
    print("\nMatching EIC names against paper authors...")
    results = find_matches(eic_names, author_forms)
    
    print("\nWriting crosswalk...")
    write_crosswalk(args.output, results)
    
    # Print some examples
    print("\n--- Sample matches ---")
    for name, matched in list(results.items())[:20]:
        if matched:
            print(f"  {name} -> {matched[:5]}{' ...' if len(matched) > 5 else ''}")
        else:
            print(f"  {name} -> NO MATCH")


if __name__ == "__main__":
    main()

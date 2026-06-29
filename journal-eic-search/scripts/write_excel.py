"""
Write journal EIC data to 期刊目录.xlsx.

Two-sheet layout:
  Sheet1 "期刊元数据": A=期刊编号 B=期刊名 C=人大分级 D=中科院分区 E=JCR分区 F=ISSN G=信息来源 H=备注
  Sheet2 "EIC_Long": A=期刊编号 B=期刊名 C=人大分级 D=ISSN E=EIC序号
                      F=姓名 G=单位 H=城市 I=国籍 J=性别 K=上任时间 L=离任时间 M=信息来源
  One row per EIC, unlimited count.

Usage:
    python write_excel.py <xlsx_path> --read
    python write_excel.py <xlsx_path> --add-journals <csv_path>
    python write_excel.py <xlsx_path> --migrate
"""
import sys
import csv
import openpyxl
from pathlib import Path

MAX_EICS = 6  # For Sheet1 backward compat only; Sheet2 has no limit
EIC_COLS = 7
ISSN_COL = 6
SOURCE_COL = 7
NOTES_COL = 8


def read_journals(xlsx_path):
    """Read journal metadata from Sheet1."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Sheet1"]
    journals = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        if not row or len(row) < 2:
            continue
        name = row[1]
        if not name or str(name).strip() == "None":
            continue
        info = {
            "row": row_idx,
            "id": str(row[0]) if row[0] else "",
            "name": str(name).strip(),
            "ruc_grade": str(row[2]).strip() if row[2] else "",
            "issn": str(row[5]).strip() if len(row) > 5 and row[5] and str(row[5]) != "None" else "",
            "source": str(row[6]).strip() if len(row) > 6 and row[6] and str(row[6]) != "None" else "",
            "notes": str(row[7]).strip() if len(row) > 7 and row[7] and str(row[7]) != "None" else "",
        }
        journals.append(info)
    return journals


def read_eics_long(xlsx_path):
    """Read all EIC records from Sheet2 EIC_Long."""
    wb = openpyxl.load_workbook(xlsx_path)
    if "EIC_Long" not in wb.sheetnames:
        return []
    ws = wb["EIC_Long"]
    eics = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row or len(row) < 6:
            continue
        name = row[1]
        if not name or str(name).strip() == "None":
            continue
        eic = {
            "journal_name": str(name).strip(),
            "journal_id": str(row[0]) if row[0] else "",
            "ruc_grade": str(row[2]).strip() if row[2] else "",
            "issn": str(row[3]).strip() if row[3] else "",
            "seq": row[4],
            "eic_name": str(row[5]).strip() if row[5] else "",
            "institution": str(row[6]).strip() if len(row) > 6 and row[6] else "",
            "city": str(row[7]).strip() if len(row) > 7 and row[7] else "",
            "nationality": str(row[8]).strip() if len(row) > 8 and row[8] else "",
            "gender": str(row[9]).strip() if len(row) > 9 and row[9] else "",
            "start_year": str(row[10]).strip() if len(row) > 10 and row[10] else "",
            "end_year": str(row[11]).strip() if len(row) > 11 and row[11] else "",
            "source": str(row[12]).strip() if len(row) > 12 and row[12] else "",
        }
        eics.append(eic)
    return eics


def init_long_sheet(xlsx_path):
    """Initialize or reset the EIC_Long sheet with headers."""
    wb = openpyxl.load_workbook(xlsx_path)
    if "EIC_Long" in wb.sheetnames:
        del wb["EIC_Long"]
    ws = wb.create_sheet("EIC_Long")
    headers = ["期刊编号", "期刊名", "人大分级", "ISSN", "EIC序号",
               "姓名", "单位", "单位所在城市", "国籍", "性别", "上任时间", "离任时间", "信息来源"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i).value = h
    wb.save(xlsx_path)


def write_eics_long(xlsx_path, journal_name, eic_list, source=""):
    """Write EIC data to EIC_Long sheet (long format). Replaces all rows for this journal."""
    wb = openpyxl.load_workbook(xlsx_path)
    
    # Ensure Sheet1 headers and EIC_Long exists
    ws1 = wb["Sheet1"]
    _write_sheet1_headers(ws1)
    
    if "EIC_Long" not in wb.sheetnames:
        init_long_sheet(xlsx_path)
        wb = openpyxl.load_workbook(xlsx_path)
    
    ws2 = wb["EIC_Long"]
    
    # Find journal metadata from Sheet1
    jid = jgrade = jissn = ""
    target_row = None
    for r in range(2, ws1.max_row + 1):
        v = ws1.cell(row=r, column=2).value
        if v and str(v).strip().upper() == journal_name.strip().upper():
            jid = str(ws1.cell(row=r, column=1).value or "")
            jgrade = str(ws1.cell(row=r, column=3).value or "")
            jissn = str(ws1.cell(row=r, column=6).value or "")
            target_row = r
            break
    
    if target_row is None:
        target_row = ws1.max_row + 1
        existing = sum(1 for r in range(2, target_row) if ws1.cell(row=r, column=2).value)
        jid = f"{existing + 1:05d}"
        ws1.cell(row=target_row, column=1).value = jid
        ws1.cell(row=target_row, column=2).value = journal_name
    
    # Write source to Sheet1
    if source:
        ws1.cell(row=target_row, column=SOURCE_COL).value = source
    
    # Remove existing rows for this journal from Sheet2
    rows_to_delete = []
    for r in range(2, ws2.max_row + 1):
        v = ws2.cell(row=r, column=2).value
        if v and str(v).strip().upper() == journal_name.strip().upper():
            rows_to_delete.append(r)
    
    # Delete from bottom up
    for r in reversed(rows_to_delete):
        ws2.delete_rows(r)
    
    # Append new rows
    next_row = ws2.max_row + 1
    # Make sure header exists
    if ws2.cell(row=1, column=1).value is None:
        headers = ["期刊编号", "期刊名", "人大分级", "ISSN", "EIC序号",
                   "姓名", "单位", "单位所在城市", "国籍", "性别", "上任时间", "离任时间", "信息来源"]
        for i, h in enumerate(headers, 1):
            ws2.cell(row=1, column=i).value = h
        next_row = 2
    
    sorted_eics = sorted(eic_list, key=lambda e: int(e.get("start_year", "0") or "0"), reverse=True)
    
    for i, eic in enumerate(sorted_eics, 1):
        r = next_row + i - 1
        ws2.cell(row=r, column=1).value = jid
        ws2.cell(row=r, column=2).value = journal_name
        ws2.cell(row=r, column=3).value = jgrade
        ws2.cell(row=r, column=4).value = jissn
        ws2.cell(row=r, column=5).value = i
        ws2.cell(row=r, column=6).value = eic.get("name", "")
        ws2.cell(row=r, column=7).value = eic.get("institution", "")
        ws2.cell(row=r, column=8).value = eic.get("city", "")
        ws2.cell(row=r, column=9).value = eic.get("nationality", "")
        ws2.cell(row=r, column=10).value = eic.get("gender", "")
        ws2.cell(row=r, column=11).value = str(eic.get("start_year", ""))
        ws2.cell(row=r, column=12).value = str(eic.get("end_year", "")) if eic.get("end_year") else ""
        ws2.cell(row=r, column=13).value = source
    
    wb.save(xlsx_path)
    print(f"Written {len(sorted_eics)} EICs for '{journal_name}' to EIC_Long")


def add_journals(xlsx_path, csv_path):
    """Add new journals from CSV to Sheet1."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Sheet1"]
    
    _write_sheet1_headers(ws)
    
    new_journals = []
    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            new_journals.append(row)
    
    existing_names = set()
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[1]:
            existing_names.add(str(row[1]).strip().upper())
    
    next_row = ws.max_row + 1
    next_id = len(existing_names) + 1
    added = 0
    
    for j in new_journals:
        title = j["journal_title"].strip()
        if title.upper() in existing_names:
            continue
        ws.cell(row=next_row, column=1).value = f"{next_id:05d}"
        ws.cell(row=next_row, column=2).value = title
        ws.cell(row=next_row, column=3).value = j["section"]
        ws.cell(row=next_row, column=4).value = None
        ws.cell(row=next_row, column=5).value = None
        ws.cell(row=next_row, column=6).value = j.get("issn", "")
        next_id += 1
        next_row += 1
        existing_names.add(title.upper())
        added += 1
    
    wb.save(xlsx_path)
    print(f"Added {added} new journals.")


def _write_sheet1_headers(ws):
    """Write Sheet1 column headers."""
    headers = {
        1: "期刊编号", 2: "期刊名", 3: "人大分级", 4: "中科院分区", 5: "JCR分区",
        6: "ISSN", 7: "信息来源", 8: "备注"
    }
    for col, val in headers.items():
        cell = ws.cell(row=1, column=col)
        if cell.value is None or str(cell.value) == "None":
            cell.value = val
        elif col >= 7:  # Always overwrite source/notes headers
            cell.value = val


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage:")
        print("  python write_excel.py <xlsx_path> --read")
        print("  python write_excel.py <xlsx_path> --add-journals <csv_path>")
        sys.exit(1)
    
    xlsx_path = sys.argv[1]
    
    if len(sys.argv) >= 3 and sys.argv[2] == "--read":
        journals = read_journals(xlsx_path)
        for j in journals:
            print(f"[Row {j['row']}] {j['name']} ({j['ruc_grade']}) | ISSN: {j['issn']} | Source: {j['source']}")
        
        eics = read_eics_long(xlsx_path)
        if eics:
            from collections import Counter
            cnt = Counter(e["journal_name"] for e in eics)
            print(f"\nEIC_Long: {len(eics)} records across {len(cnt)} journals")
            for jn, c in sorted(cnt.items()):
                print(f"  {c} EICs: {jn}")
    elif len(sys.argv) >= 4 and sys.argv[2] == "--add-journals":
        add_journals(xlsx_path, sys.argv[3])

"""
Post-process raw WoS crawler output:
  1. Fix country names -> ISO 3166-1 alpha-3
  2. Fix city extraction (handle England-as-city bug, suburb-vs-city, multi-campus)
  3. Convert to long-format Excel (Papers + Authors sheets)

Usage:
    python postprocess.py --input <raw_csv> --output <xlsx> --mode {fix|to-long|all}
"""
import sys
import csv
import re
import openpyxl
from pathlib import Path

# ── Country map (WoS names -> ISO 3166-1 alpha-3) ─────────────
COUNTRY_MAP = {
    "United States": "USA", "United States of America": "USA", "USA": "USA", "US": "USA",
    "England": "England", "Scotland": "Scotland", "Wales": "Wales", 
    "Northern Ireland": "Northern Ireland",
    "United Kingdom": "UK", "UK": "UK", "GBR": "UK",
    "Germany": "DEU", "France": "FRA", "Italy": "ITA", "Spain": "ESP",
    "Netherlands": "NLD", "Belgium": "BEL", "Switzerland": "CHE",
    "Sweden": "SWE", "Norway": "NOR", "Denmark": "DNK", "Finland": "FIN",
    "Canada": "CAN", "Australia": "AUS", "New Zealand": "NZL",
    "Japan": "JPN", "China": "CHN", "PR China": "CHN", "Peoples R China": "CHN",
    "South Korea": "KOR", "Republic of Korea": "KOR",
    "Singapore": "SGP", "Hong Kong": "HKG", "Taiwan": "TWN",
    "India": "IND", "Israel": "ISR", "Russia": "RUS", "Brazil": "BRA",
    "Turkey": "TUR", "South Africa": "ZAF", "Portugal": "PRT", "Austria": "AUT",
    "Ireland": "IRL", "Greece": "GRC", "Poland": "POL",
    "Czech Republic": "CZE", "Hungary": "HUN", "Romania": "ROU",
    "United Arab Emirates": "ARE", "Saudi Arabia": "SAU",
    "Luxembourg": "LUX", "Cyprus": "CYP", "Colombia": "COL",
    "Pakistan": "PAK", "Philippines": "PHL", "Malaysia": "MYS",
    "Thailand": "THA", "Vietnam": "VNM", "Indonesia": "IDN",
    "Egypt": "EGY", "Nigeria": "NGA", "Kenya": "KEN",
    "Peru": "PER", "Uruguay": "URY", "Ecuador": "ECU", "Chile": "CHL",
    "Argentina": "ARG", "Mexico": "MEX",
    "Croatia": "HRV", "Slovenia": "SVN", "Slovakia": "SVK",
    "Estonia": "EST", "Latvia": "LVA", "Lithuania": "LTU",
    "Iceland": "ISL", "Bulgaria": "BGR", "Norway": "NOR",
}

# Countries whose sub-regions commonly appear as address components
UK_REGIONS = {"GREAT BRITAIN", "BRITAIN"}  # England/Scotland/Wales are now valid countries
CITY_STATE_CITIES = {"SINGAPORE", "MONACO", "VATICAN CITY", "HONG KONG", "MACAU", "GIBRALTAR"}

# Well-known UK cities → country = England (for WoS convention)
UK_CITIES = {"LONDON", "CAMBRIDGE", "OXFORD", "MANCHESTER", "BIRMINGHAM", "BRISTOL",
             "LEEDS", "LIVERPOOL", "SHEFFIELD", "NOTTINGHAM", "SOUTHAMPTON",
             "NEWCASTLE UPON TYNE", "NEWCASTLE", "LEICESTER", "COVENTRY", "WARWICK",
             "BRIGHTON", "NORWICH", "EXETER", "YORK", "BATH", "DURHAM", "CARDIFF",
             "EDINBURGH", "GLASGOW", "ABERDEEN", "BELFAST", "READING", "ESSEX",
             "SURREY", "KENT", "SUSSEX", "LANCASTER", "LOUGHBOROUGH", "CRANFIELD",
              "BRUNEL", "UXBRIDGE", "CANTERBURY", "COLCHESTER", "GUILDFORD"}

# Canadian FSA (Forward Sortation Area) -> City mapping
FSA_MAP = {
    "M5S": "Toronto", "M5V": "Toronto", "M3J": "Toronto", "M5G": "Toronto",
    "M5T": "Toronto", "M5R": "Toronto",
    "N6A": "London", "N6C": "London",
    "K1A": "Ottawa", "K1S": "Ottawa", "K1N": "Ottawa", "K1G": "Ottawa",
    "L5L": "Mississauga", "K7L": "Kingston", "L8S": "Hamilton",
    "N2L": "Waterloo", "L2S": "St. Catharines",
    "V6T": "Vancouver", "V6R": "Vancouver", "V6H": "Vancouver",
    "V5A": "Burnaby", "V5Z": "Vancouver", "V8W": "Victoria",
    "T2N": "Calgary", "T2P": "Calgary", "T6G": "Edmonton",
    "R3T": "Winnipeg", "B3H": "Halifax",
    "H3B": "Montreal", "H3C": "Montreal", "H3G": "Montreal",
    "H2X": "Montreal", "H2Y": "Montreal",
}
CAN_POSTAL_RE = re.compile(r'^([A-Za-z]\d[A-Za-z]) \d[A-Za-z]\d$')
MONTREAL_KW = {'montreal', 'mcgill', 'hec montreal', 'hec', 'udem', 'uqam',
               'concordia', 'cirano', 'univ quebec montreal'}
QUEBEC_KW = {'univ laval', 'quebec city', 'universite laval'}
OTTAWA_KW = {'bank canada', 'bank of canada', 'competit bur canada'}

# Country prefix + postal code embedded in city
CODE_PREFIX_RE = re.compile(r'^([A-Z]{1,2}[-\s])\d{4,8}[-\s]?[A-Z]{0,2}\s*')

# ── Systematic cleaning patterns ────────────────────────────────────────────
# Institution: strip leading "N \n" (WoS raw format: "1 \nHarvard Univ")
INST_NUM_RE = re.compile(r'^\d+\s*\n\s*')
# City: strip postal code suffix like " Barcelona 08005" or "Kyoto 6068501"
CITY_SUFFIX_RE = re.compile(r'\s+\d{4,8}$')
# City: pure numeric (postal code only, no city name)
CITY_NUMERIC_ONLY_RE = re.compile(r'^\d+$')

# When city field actually contains a country name
CITY_AS_COUNTRY = {
    "U Arab Emirates": "ARE", "Turkiye": "TUR", "Bangladesh": "BGD",
    "DEM REP CONGO": "COD", "Sri Lanka": "LKA", "Uganda": "UGA",
    "Bahrain": "BHR", "Kiev": "UKR", "Costa Rica": "CRI",
    "Ethiopia": "ETH", "Afghanistan": "AFG", "Ghana": "GHA",
    "Serbia": "SRB", "Cote Ivoire": "CIV", "Bolivia": "BOL",
    "Dominican Rep": "DOM", "Benin": "BEN", "Qatar": "QAT",
    "Venezuela": "VEN", "Mongolia": "MNG", "Guyana": "GUY", "Bermuda": "BMU",
    "Iran": "IRN", "BELARUS": "BLR", "Panama": "PAN",
    "Jordan": "JOR", "Lesotho": "LSO",
}
COUNTRY_TO_CAPITAL = {
    "ARE": "Abu Dhabi", "TUR": "Ankara", "BGD": "Dhaka", "COD": "Kinshasa",
    "LKA": "Colombo", "UGA": "Kampala", "BHR": "Manama", "UKR": "Kiev",
    "CRI": "San Jose", "ETH": "Addis Ababa", "AFG": "Kabul", "GHA": "Accra",
    "SRB": "Belgrade", "CIV": "Abidjan", "BOL": "La Paz", "DOM": "Santo Domingo",
    "BEN": "Cotonou", "QAT": "Doha", "VEN": "Caracas", "MNG": "Ulaanbaatar",
    "GUY": "Georgetown", "BMU": "Hamilton",
    "IRN": "Tehran", "BLR": "Minsk", "PAN": "Panama City",
    "JOR": "Amman", "LSO": "Maseru",
}

# Institution -> (City, Country) lookup for city inference
INST_CITY_MAP = {
    "harvard": ("Cambridge", "USA"), "mit": ("Cambridge", "USA"),
    "stanford": ("Palo Alto", "USA"), "yale": ("New Haven", "USA"),
    "princeton": ("Princeton", "USA"), "princeton univ": ("Princeton", "USA"),
    "uchicago": ("Chicago", "USA"),
    "columbia univ": ("New York", "USA"), "nyu": ("New York", "USA"),
    "berkeley": ("Berkeley", "USA"), "ucla": ("Los Angeles", "USA"),
    "univ michigan": ("Ann Arbor", "USA"), "duke": ("Durham", "USA"),
    "johns hopkins": ("Baltimore", "USA"), "northwestern": ("Evanston", "USA"),
    "upenn": ("Philadelphia", "USA"), "univ penn": ("Philadelphia", "USA"),
    "lse": ("London", "GBR"), "oxford": ("Oxford", "GBR"),
    "cambridge univ": ("Cambridge", "GBR"), "cornell": ("Ithaca", "USA"),
    "univ maryland": ("College Park", "USA"), "penn state": ("State College", "USA"),
    "penn state univ": ("State College", "USA"),
    "texas a&m": ("College Station", "USA"), "texas a&m univ": ("College Station", "USA"),
    "arizona state": ("Tempe", "USA"),
    "carnegie mellon": ("Pittsburgh", "USA"), "emory": ("Atlanta", "USA"),
    "georgetown": ("Washington", "USA"), "george washington univ": ("Washington", "USA"),
    "indiana univ": ("Bloomington", "USA"), "michigan state": ("East Lansing", "USA"),
    "ohio state": ("Columbus", "USA"), "purdue": ("West Lafayette", "USA"),
    "rice univ": ("Houston", "USA"), "univ arizona": ("Tucson", "USA"),
    "uc davis": ("Davis", "USA"), "uc irvine": ("Irvine", "USA"),
    "uc santa barbara": ("Santa Barbara", "USA"), "uc san diego": ("San Diego", "USA"),
    "univ colorado": ("Boulder", "USA"), "univ florida": ("Gainesville", "USA"),
    "univ georgia": ("Athens", "USA"), "univ illinois": ("Champaign", "USA"),
    "univ iowa": ("Iowa City", "USA"), "univ notre dame": ("South Bend", "USA"),
    "univ rochester": ("Rochester", "USA"), "univ southern california": ("Los Angeles", "USA"),
    "univ texas": ("Austin", "USA"), "univ utah": ("Salt Lake City", "USA"),
    "univ virginia": ("Charlottesville", "USA"), "univ washington": ("Seattle", "USA"),
    "vanderbilt": ("Nashville", "USA"), "washington univ": ("St Louis", "USA"),
    "univ north carolina": ("Chapel Hill", "USA"), "univ alberta": ("Edmonton", "CAN"),
    "univ montreal": ("Montreal", "CAN"), "mcgill": ("Montreal", "CAN"),
    "univ toronto": ("Toronto", "CAN"), "univ british columbia": ("Vancouver", "CAN"),
    "univ western ontario": ("London", "CAN"), "queens univ": ("Kingston", "CAN"),
    "univ zurich": ("Zurich", "CHE"), "eth zurich": ("Zurich", "CHE"),
    "pompeu fabra": ("Barcelona", "ESP"), "univ barcelona": ("Barcelona", "ESP"),
    "bocconi": ("Milan", "ITA"), "goethe univ": ("Frankfurt", "DEU"),
    "univ bonn": ("Bonn", "DEU"), "univ mannheim": ("Mannheim", "DEU"),
    "univ munich": ("Munich", "DEU"), "sciences po": ("Paris", "FRA"),
    "paris sch econ": ("Paris", "FRA"), "toulouse sch econ": ("Toulouse", "FRA"),
    "univ oslo": ("Oslo", "NOR"), "stockholm univ": ("Stockholm", "SWE"),
    "stockholm sch econ": ("Stockholm", "SWE"), "ku leuven": ("Leuven", "BEL"),
    "univ melbourne": ("Melbourne", "AUS"), "univ sydney": ("Sydney", "AUS"),
    "australian natl univ": ("Canberra", "AUS"), "monash": ("Melbourne", "AUS"),
    "monash univ": ("Melbourne", "AUS"),
    "univ new s wales": ("Sydney", "AUS"), "queensland univ": ("Brisbane", "AUS"),
    "univ queensland": ("Brisbane", "AUS"),
    "univ technol sydney": ("Sydney", "AUS"), "univ adelaide": ("Adelaide", "AUS"),
    "univ western australia": ("Perth", "AUS"),
    "natl univ singapore": ("Singapore", "SGP"),
    "univ hong kong": ("Hong Kong", "HKG"),
    "peking univ": ("Beijing", "CHN"), "tsinghua": ("Beijing", "CHN"),
    "univ tokyo": ("Tokyo", "JPN"), "keio univ": ("Tokyo", "JPN"),
    "eief": ("Rome", "ITA"),
    # Additional
    "queensland univ technol": ("Brisbane", "AUS"),
    "univ s australia": ("Adelaide", "AUS"),
    "univ wollongong": ("Wollongong", "AUS"),
    "inst tecnol autonomo mexico": ("Mexico City", "MEX"),
    "melbourne business sch": ("Melbourne", "AUS"),
    "royal brisbane & womens hosp": ("Brisbane", "AUS"),
    "univ reims": ("Reims", "FRA"),
    "resources future inc": ("Washington", "USA"),
    "shahrood univ technol": ("Shahrood", "IRN"),
    "univ tehran med sci": ("Tehran", "IRN"),
    "belarusian state univ": ("Minsk", "BLR"),
    "gorgas mem inst hlth studies": ("Panama City", "PAN"),
    "jordan univ sci & technol": ("Irbid", "JOR"),
}

CITY_VARIANTS = {
    "Tel Aviv-Yafo": "Tel Aviv", "Tel-Aviv": "Tel Aviv",
    "Washington, DC": "Washington", "Mexico City, DF": "Mexico City",
    "Ciudad de Mexico": "Mexico City", "CDMX": "Mexico City",
    "St. Louis": "St Louis", "Berkeley, CA": "Berkeley",
}

# US/Canada/Australia state abbreviations (commonly appear in addresses)
STATE_ABBR = {"AL","AK","AZ","AR","CA","CO","CT","DE","DC","FL","GA","HI","ID","IL","IN","IA",
              "KS","KY","LA","ME","MD","MA","MI","MN","MS","MO","MT","NE","NV","NH","NJ","NM",
              "NY","NC","ND","OH","OK","OR","PA","RI","SC","SD","TN","TX","UT","VT","VA","WA",
              "WV","WI","WY","ON","QC","BC","AB","MB","SK","NS","NB","NL","PE","NT","YT","NU",
              "NSW","VIC","QLD","WA","SA","TAS","ACT","NT"}

INST_KEYWORDS = {'univ', 'college', 'school', 'sch', 'inst', 'dept', 'department',
                 'laboratory', 'lab', 'center', 'centre', 'hospital', 'corp',
                 'inc', 'ltd', 'co', 'llc', 'gmbh', 'bv', 'sa', 'ag', 'found',
                 'fac', 'fed', 'polytech', 'econ', 'management', 'business',
                 'acad', 'assoc', 'soc', 'minist', 'bureau', 'comm', 'authority',
                 'council', 'board', 'off', 'div', 'sect', 'branch', 'program',
                 'project', 'initiative', 'network', 'consortium', 'alliance'}

ZIP_RE = re.compile(r'^\d{4,6}(?:-\d+)?$')


def normalize_country(name):
    """WoS country name -> ISO 3166-1 alpha-3."""
    if not name:
        return ""
    name = name.strip()
    for key, val in COUNTRY_MAP.items():
        if key.lower() == name.lower():
            return val
    return name


def _is_institution_part(part):
    """Check if part is an institutional name, not a city."""
    p = part.strip().lower()
    # Match keywords as whole words or at word boundaries
    words = set(p.replace('-', ' ').replace('/', ' ').split())
    for kw in INST_KEYWORDS:
        if kw in words:
            return True
    # Also check if part contains institution keyword as prefix (e.g. "ColumbiaUniv")
    for kw in INST_KEYWORDS:
        if len(kw) >= 4 and p.startswith(kw):
            return True
    return False


def extract_city_country(affiliation, raw_country=""):
    """
    Extract city and country from an affiliation string.
    
    Handles cases like:
    - "IGC, London, England" -> city="London", country="GBR"
    - "Univ Chicago, Chicago, IL 60637 USA" -> city="Chicago", country="USA"
    - "Ecole Polytech Fed Lausanne, Lausanne, Switzerland" -> city="Lausanne", country="CHE"
    """
    if not affiliation:
        return "", ""
    
    affiliation = affiliation.strip()
    parts = [p.strip() for p in affiliation.split(",")]
    
    if not parts:
        return "", ""
    
    # First pass: identify country from any part (exact or contained)
    country = ""
    for i in range(len(parts) - 1, -1, -1):
        p = parts[i]
        p_upper = p.strip().upper()
        # Try exact match first
        matched = False
        for cname, ccode in COUNTRY_MAP.items():
            cu = cname.upper()
            if p_upper == cu or p_upper == ccode.upper():
                country = ccode
                matched = True
                break
            # Handle embedded country like "IL 60637 USA" -> USA
            if len(cu) >= 3 and cu in p_upper.split():
                country = ccode
                matched = True
                break
        if matched:
            break
    
    # If no country found in parts, use raw_country
    if not country and raw_country:
        country = normalize_country(raw_country)
    
    # Second pass: identify city from parts
    # Strategy: scan from end, skip known non-city parts (zip, state, country)
    candidate_parts = list(parts)
    
    # Remove zip codes
    candidate_parts = [p for p in candidate_parts if not ZIP_RE.match(p.strip())]
    
    # Remove country part (if found)
    for cname, _ in COUNTRY_MAP.items():
        if candidate_parts and candidate_parts[-1].strip().upper() == cname.upper():
            candidate_parts = candidate_parts[:-1]
            break
    
    # Remove state abbreviation from end (handles "IL 60637 USA" -> removes IL portion)
    if candidate_parts:
        last = candidate_parts[-1].strip().upper()
        for abbr in STATE_ABBR:
            if last.startswith(abbr + " ") or last == abbr:
                # Remove the state abbreviation, keep the rest
                if last == abbr:
                    candidate_parts = candidate_parts[:-1]
                else:
                    candidate_parts[-1] = last[len(abbr):].strip()
                break
    
    # Now find city: scan remaining parts from end for a plausible city name
    city = ""
    for i in range(len(candidate_parts) - 1, -1, -1):
        p = candidate_parts[i].strip()
        p_upper = p.upper()
        
        # Skip zip-like patterns (containing numbers)
        if re.search(r'\d{5}', p):
            continue
        # Skip known UK regions being used wrongly as city
        if p_upper in UK_REGIONS:
            continue
        # Skip institutional parts (keyword-based)
        if _is_institution_part(p):
            continue
        # Skip state abbreviations
        if p_upper in STATE_ABBR:
            continue
        # Skip country parts (full country names that weren't caught earlier)
        country_match = False
        for cname in COUNTRY_MAP:
            if cname.upper() == p_upper or p_upper.startswith(cname.upper()):
                country_match = True
                break
        if country_match:
            continue
        
        # This looks like a city
        city = p
        break
    
    # Fallback: if no city found and affiliation has enough parts
    if not city and len(parts) >= 2:
        # Try to extract city from between institution and state/country
        for i in range(1, len(parts)):
            p = parts[i].strip()
            p_upper = p.upper()
            if p_upper in UK_REGIONS or p_upper in {c.upper() for c in COUNTRY_MAP.values()}:
                continue
            if p_upper in STATE_ABBR:
                continue
            if ZIP_RE.match(p):
                continue
            if any(kw in p.lower() for kw in INST_KEYWORDS):
                continue
            city = p
            break
    
    # City-state handling
    if country == "SGP":
        city = "Singapore"
    if country == "MCO":
        city = "Monaco"
    if country == "HKG":
        city = "Hong Kong"
    
    return city, country


def process_raw_csv(input_path, output_path):
    """Fix locations in raw crawler CSV."""
    rows = []
    with open(input_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        for row in reader:
            rows.append(row)
    
    print(f"Read {len(rows)} papers")
    fixed_count = 0

    for row in rows:
        # Fix addresses: parse Authors field and Addresses field
        # The raw crawler stores addresses as " || " separated strings
        
        addresses_raw = row.get("Addresses", "")
        if addresses_raw:
            addr_list = [a.strip() for a in addresses_raw.split("||") if a.strip()]
            for i, addr in enumerate(addr_list[:10], 1):
                city, country = extract_city_country(addr)
                if city or country:
                    row[f"Addr{i}_City"] = city
                    row[f"Addr{i}_Country"] = country
                    fixed_count += 1
        
        # Fix corresponding author address
        corr_addr = row.get("Corresponding Address", "")
        if corr_addr:
            city, country = extract_city_country(corr_addr)
            if city:
                row["Corr_City"] = city
            if country:
                row["Corr_Country"] = country

    # Write back
    all_fieldnames = list(fieldnames)
    for new_col in [f"Addr{i}_{s}" for i in range(1, 11) for s in ["City","Country"]] + ["Corr_City", "Corr_Country"]:
        if new_col not in all_fieldnames:
            all_fieldnames.append(new_col)

    with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=all_fieldnames, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(rows)
    
    print(f"Fixed {fixed_count} city/country entries -> {output_path}")


def _clean_inst(inst):
    """Strip leading 'N \\n' from institution name (WoS raw format)."""
    if not inst:
        return ""
    inst = str(inst).strip()
    inst = INST_NUM_RE.sub("", inst)
    return inst

def _clean_city(city, country=""):
    """Strip postal codes from city name and apply variant normalization."""
    if not city:
        return ""
    city = str(city).strip()
    # Pure numeric (postal code only, no city name) -> empty
    if CITY_NUMERIC_ONLY_RE.match(city):
        return ""
    # Remove code prefix like "CH-8006 Zurich" -> "Zurich"
    city = CODE_PREFIX_RE.sub("", city)
    # Remove code suffix like "Barcelona 08005" -> "Barcelona", "Kyoto 6068501" -> "Kyoto"
    city = CITY_SUFFIX_RE.sub("", city)
    # Apply variants
    city = CITY_VARIANTS.get(city, city)
    return city

def _infer_city_country_from_inst(inst):
    """Given cleaned institution name, return (city, country) or (None, None)."""
    if not inst:
        return None, None
    inst_lower = _clean_inst(inst).lower().strip()
    if inst_lower in INST_CITY_MAP:
        return INST_CITY_MAP[inst_lower]
    return None, None

def _infer_country_from_city(city):
    """If city field is actually a country name, return ISO code."""
    if not city:
        return None
    city_str = str(city).strip()
    if city_str in CITY_AS_COUNTRY:
        iso = CITY_AS_COUNTRY[city_str]
        capital = COUNTRY_TO_CAPITAL.get(iso, city_str)
        return iso, capital
    return None, None


def convert_to_long(input_csv, output_xlsx):
    """Convert wide-format CSV to long-format Excel (Papers + Authors sheets).
    Auto-detects old format (Author 1-10 columns) vs new format (Authors pipe-separated).
    """
    with open(input_csv, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        fieldnames = list(reader.fieldnames)
        
        # Detect format
        is_new_format = "Authors" in fieldnames and "Addresses" in fieldnames
        print(f"Format: {'new (pipe-separated)' if is_new_format else 'old (Author 1-10 columns)'}")
        
        papers = []
        authors = []
        pid_counter = 1

        for row in reader:
            wos_id = row.get("WoS_ID", row.get("WoS ID", f"P{pid_counter:06d}"))
            
            paper = {
                "paper_id": wos_id,
                "journal_name": row.get("Source_Journal", row.get("Journal_Name", "")),
                "article_title": row.get("Article_Title", row.get("Article Title", "")),
                "source_journal": row.get("Source_Journal", row.get("Source Journal", "")),
                "published_date": row.get("Published_Date", row.get("Published Date", "")),
                "early_access": row.get("Early_Access", row.get("Early Access", "")),
                "keywords": row.get("Keywords", ""),
                "wos_categories": row.get("WoS_Categories", row.get("WoS Categories", "")),
                "document_type": (row.get("Document_Type", row.get("Document Type", "")) or "").rstrip(";"),
                "citations_wos_core": row.get("Citations_WoS_Core",
                                        row.get("Citation Count",
                                        row.get("Citations WoS Core", "0"))),
                "citations_all_db": row.get("Citations_All_DB",
                                      row.get("Citations All DB",
                                      row.get("Times Cited All DB", "0"))),
                "cited_references": row.get("Cited_References",
                                       row.get("Cited References", "0")),
            }
            papers.append(paper)

            corr_author = row.get("Corr_Author", row.get("Corresponding Author", ""))

            if is_new_format:
                # New format: pipe-separated strings
                authors_str = row.get("Authors", "")
                author_list = [a.strip() for a in authors_str.split("|") if a.strip()]
                addresses_str = row.get("Addresses", "")
                addr_list = [a.strip() for a in addresses_str.split("||") if a.strip()]
                
                for i, aname in enumerate(author_list):
                    author = {
                        "paper_id": wos_id, "author_seq": i+1, "author_name": aname,
                        "is_corresponding": "TRUE" if aname == corr_author else "FALSE",
                        "institution": "", "city": "", "country": "",
                    }
                    if i < len(addr_list):
                        city, country = extract_city_country(addr_list[i])
                        author["city"] = city
                        author["country"] = country
                        if not author["country"] and author["city"].upper() in UK_CITIES:
                            author["country"] = "England"
                        inst_parts = [p.strip() for p in addr_list[i].split(",")]
                        author["institution"] = inst_parts[0] if inst_parts else ""
                    # ── Apply systematic cleaning ──
                    author["institution"] = _clean_inst(author["institution"])
                    author["city"] = _clean_city(author["city"], author["country"])
                    # Infer missing city from institution
                    if not author["city"]:
                        inf_city, inf_country = _infer_city_country_from_inst(author["institution"])
                        if inf_city:
                            author["city"] = inf_city
                            if not author["country"]:
                                author["country"] = inf_country
                    # Infer missing country from city (when city is actually country name)
                    if not author["country"] and author["city"]:
                        inf_iso, inf_capital = _infer_country_from_city(author["city"])
                        if inf_iso:
                            author["country"] = inf_iso
                            author["city"] = inf_capital
                    authors.append(author)
            else:
                # Old format: Author 1..10 columns + Address 1..3 columns
                author_list = []
                for i in range(1, 51):
                    a = row.get(f"Author {i}", "")
                    if not a or a.strip() == "None":
                        break
                    author_list.append(a.strip())
                
                for i, aname in enumerate(author_list):
                    addr_aff = row.get(f"Address {i+1} Affiliation", "")
                    addr_city = row.get(f"Address {i+1} City", "")
                    addr_country = row.get(f"Address {i+1} Country", "")
                    corr_city = row.get("Corresponding City", "")
                    corr_country = row.get("Corresponding Country", "")
                    corr_aff = row.get("Corresponding Affiliation", "")
                    
                    city = addr_city
                    country = addr_country
                    inst = addr_aff
                    
                    # Use corrected city/country if available
                    if addr_aff:
                        est_city, est_country = extract_city_country(addr_aff)
                        if est_city and (not addr_city or addr_city.lower() in ('england','scotland','wales')):
                            city = est_city
                        # Only override country if original is empty AND extraction found one
                        if not country and est_country:
                            country = est_country
                    
                    # For corresponding author, use corresponding address fields
                    if aname.lower() == corr_author.lower():
                        if corr_aff:
                            est_city, est_country = extract_city_country(corr_aff)
                            if est_city:
                                city = est_city
                            if not country and est_country:
                                country = est_country
                            inst = corr_aff
                    
                    # Fallback: UK city without country -> England
                    if not country and city.upper() in UK_CITIES:
                        country = "England"

                    # ── Apply systematic cleaning ──
                    inst = _clean_inst(inst) if inst else ""
                    city = _clean_city(city, country)
                    if not city:
                        inf_city, inf_country = _infer_city_country_from_inst(inst)
                        if inf_city:
                            city = inf_city
                            if not country:
                                country = inf_country
                    if not country and city:
                        inf_iso, inf_capital = _infer_country_from_city(city)
                        if inf_iso:
                            country = inf_iso
                            city = inf_capital

                    authors.append({
                        "paper_id": wos_id, "author_seq": i+1, "author_name": aname,
                        "is_corresponding": "TRUE" if aname.lower() == corr_author.lower() else "FALSE",
                        "institution": inst if inst and inst != "None" else "",
                        "city": city if city and city != "None" else "",
                        "country": country if country and country != "None" else "",
                    })
            
            pid_counter += 1

    # Write Excel
    import os
    os.makedirs(os.path.dirname(output_xlsx) or ".", exist_ok=True)
    wb = openpyxl.Workbook()

    # Papers sheet
    ws1 = wb.active
    ws1.title = "Papers"
    paper_headers = ["paper_id", "journal_name", "article_title", "source_journal",
                     "published_date", "early_access", "keywords", "wos_categories",
                     "document_type", "citations_wos_core", "citations_all_db", "cited_references"]
    for c, h in enumerate(paper_headers, 1):
        ws1.cell(row=1, column=c).value = h
    for r, p in enumerate(papers, 2):
        for c, h in enumerate(paper_headers, 1):
            val = p.get(h, "")
            ws1.cell(row=r, column=c).value = val

    # Authors sheet
    ws2 = wb.create_sheet("Authors")
    author_headers = ["paper_id", "author_seq", "author_name", "is_corresponding",
                      "institution", "city", "country", "remarks"]
    for c, h in enumerate(author_headers, 1):
        ws2.cell(row=1, column=c).value = h
    for r, a in enumerate(authors, 2):
        for c, h in enumerate(author_headers, 1):
            ws2.cell(row=r, column=c).value = a.get(h, "")

    wb.save(output_xlsx)
    print(f"Converted {len(papers)} papers, {len(authors)} authors -> {output_xlsx}")


if __name__ == "__main__":
    if len(sys.argv) < 4:
        print("Usage: python postprocess.py --input <raw_csv> --output <long_xlsx>")
        sys.exit(1)
    
    args = {sys.argv[i]: sys.argv[i+1] for i in range(1, len(sys.argv), 2)}
    input_file = args.get("--input", "")
    output_file = args.get("--output", "")
    
    if not output_file.endswith(".xlsx"):
        output_file = output_file.replace(".csv", ".xlsx")
    
    # Single step: raw CSV -> city/country fix -> long-format Excel
    print(f"Processing: {input_file} -> {output_file}")
    process_raw_csv(input_file, output_file.replace(".xlsx", "_temp.csv"))
    convert_to_long(output_file.replace(".xlsx", "_temp.csv"), output_file)
    # Clean up temp
    import os
    temp = output_file.replace(".xlsx", "_temp.csv")
    if os.path.exists(temp):
        os.remove(temp)
    print(f"Done: {output_file}")
